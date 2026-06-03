"""
build.py — regenerates main.js config blocks from sensor_configuration.xlsx

This script rewrites these blocks in main.js based on the Excel file:
  - SENSOR_TYPES     ← sheet "sensors" + "params"
  - PORT_TIPS, PORTS ← sheet "ports"
  - VIZ_OPTIONS      ← sheet "viz_options"
  - SURVEY_QUESTIONS ← sheet "survey_questions"

It does NOT touch the TEMPLATES block — that contains the actual Arduino
code logic and must be edited by hand in main.js.

IMPORTANT: For Excel edits to flow through to generated code, the output names
in your "sensors" sheet (e.g. "TAW", "Raw value") must match keys in
TEMPLATES.outputs (case-insensitive). If you add a new output in Excel, add a
matching entry in TEMPLATES.outputs in main.js.

Same rule applies for new sensor types (TEMPLATES.sensors) and new viz
options (TEMPLATES.viz).
"""

import os
import re
from openpyxl import load_workbook

EXCEL_FILE    = "sensor_configuration.xlsx"
MAIN_JS_FILE  = "main.js"
TEMPLATES_DIR = "templates"

wb = load_workbook(EXCEL_FILE, data_only=True)

def sheet_rows(sheet_name, skip=1):
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=skip + 1, values_only=True):
        if any(c is not None for c in row):
            rows.append([str(c).strip() if c is not None else "" for c in row])
    return rows

def safe_js_key(key):
    return re.sub(r'[^a-zA-Z0-9_]', '_', key)

sensors_raw = sheet_rows("sensors")

active_keys = []
sensor_meta = {}
outputs_map = {}

current_key = ""
for row in sensors_raw:
    key      = row[0]
    out_val  = row[1]
    out_full = row[2]
    out_tip  = row[3]
    inp_a    = row[4]
    inp_b    = row[5]
    inp_c    = row[6]
    inp_d    = row[7]
    inp_e    = row[8]
    equation = row[9] if len(row) > 9 else ""

    if key:
        current_key = safe_js_key(key)
        if current_key not in sensor_meta:
            sensor_meta[current_key] = {
                "label":   key,
                "tip":     out_tip,
                "raw_key": key,
            }
            active_keys.append(current_key)

    if not current_key:
        continue

    if out_val and out_val.lower() != "for all params, col min max default":
        outputs_map.setdefault(current_key, [])
        if not any(o["value"] == out_val for o in outputs_map[current_key]):
            outputs_map[current_key].append({
                "value":    out_val,
                "tip":      out_tip,
                "full":     out_full,
                "equation": equation,
                "inputs":   [inp_a, inp_b, inp_c, inp_d, inp_e],
            })

params_raw = sheet_rows("params")
params_map = {}

for row in params_raw:
    key       = safe_js_key(row[0])
    p_name    = row[1]
    # Support both 7-col (with display_name) and 6-col (without) structures
    # col 0: sensor_key, 1: param_name, then either:
    # 7-col: 2=display, 3=label, 4=min, 5=max, 6=default, 7=units
    # 6-col: 2=label,   3=min,   4=max, 5=default, 6=units
    if len(row) >= 8 and row[2] and not str(row[2]).replace(".", "").replace("-", "").lstrip().isdigit():
        p_display = row[2]
        p_label   = row[3]
        p_min     = row[4]
        p_max     = row[5]
        p_val     = row[6] if len(row) > 6 else "0"
        p_units   = row[7] if len(row) > 7 else ""
    else:
        p_display = ""
        p_label   = row[2]
        p_min     = row[3]
        p_max     = row[4]
        p_val     = row[5] if len(row) > 5 else "0"
        p_units   = row[6] if len(row) > 6 else ""

    if not key or not p_name:
        continue

    params_map.setdefault(key, [])
    params_map[key].append({
        "name":    p_name,
        "display": p_display if p_display else p_name,
        "label":   p_label,
        "min":     p_min,
        "max":     p_max,
        "value":   str(p_val) if p_val != "" else "0",
        "units":   str(p_units) if p_units else "",
    })

viz_raw      = sheet_rows("viz_options")
active_viz   = [r for r in viz_raw if r[3].upper() == "TRUE"]
survey_raw   = sheet_rows("survey_questions")
ports_raw    = sheet_rows("ports")
active_ports = [r[0] for r in ports_raw if r[0]]
port_tips    = {r[0]: r[1] for r in ports_raw if r[0]}


def jstr(s):
    return (s or "").replace("\\", "\\\\").replace('"', '\\"')


def build_sensor_types():
    lines = ["const SENSOR_TYPES = {"]
    for key in active_keys:
        meta    = sensor_meta[key]
        outputs = outputs_map.get(key, [])
        params  = params_map.get(key, [])
        lines.append(f'  {key}: {{')
        lines.append(f'    label: "{jstr(meta["label"])}",')
        lines.append(f'    tip: "{jstr(meta["tip"])}",')
        lines.append('    outputs: [')
        for o in outputs:
            lines.append(f'      {{')
            lines.append(f'        value: "{jstr(o["value"])}",')
            lines.append(f'        tip: "{jstr(o["tip"])}",')
            lines.append(f'      }},')
        lines.append('    ],')
        lines.append('    params: [')
        for p in params:
            lines.append(f'      {{')
            lines.append(f'        name: "{jstr(p["name"])}",')
            lines.append(f'        display: "{jstr(p["display"])}",')
            lines.append(f'        label: "{jstr(p["label"])}",')
            lines.append(f'        value: "{jstr(p["value"])}",')
            lines.append(f'        min: "{jstr(p["min"])}",')
            lines.append(f'        max: "{jstr(p["max"])}",')
            lines.append(f'        units: "{jstr(p["units"])}",')
            lines.append(f'      }},')
        lines.append('    ],')
        lines.append('  },')
    lines.append('};')
    return '\n'.join(lines)


def build_port_tips():
    lines = ["const PORT_TIPS = {"]
    for port, tip in port_tips.items():
        lines.append(f'  {port}: "{jstr(tip)}",')
    lines.append('};')
    return '\n'.join(lines)


def build_ports():
    items = ',\n'.join(f'  "{p}"' for p in active_ports)
    return f'const PORTS = [\n{items},\n];'


def build_viz_options():
    lines = ["const VIZ_OPTIONS = ["]
    for row in active_viz:
        key, label, tip = row[0], row[1], row[2]
        lines.append(f'  {{')
        lines.append(f'    value: "{jstr(key)}",')
        lines.append(f'    label: "{jstr(label)}",')
        lines.append(f'    tip: "{jstr(tip)}",')
        lines.append(f'  }},')
    lines.append('];')
    return '\n'.join(lines)


def build_survey_questions():
    lines = ["const SURVEY_QUESTIONS = ["]
    for row in survey_raw:
        key, label, qtype = row[0], row[1], row[2]
        required = "true" if row[3].upper() == "TRUE" else "false"
        extra    = row[4] if len(row) > 4 else ""
        lines.append(f'  {{')
        lines.append(f'    key: "{key}",')
        lines.append(f'    label: "{jstr(label)}",')
        lines.append(f'    type: "{qtype}",')
        lines.append(f'    required: {required},')
        if qtype.lower() == "select":
            options = [o.strip() for o in extra.split("|") if o.strip()]
            opts_js = "[" + ", ".join(f'"{jstr(o)}"' for o in options) + "]"
            lines.append(f'    options: {opts_js},')
        else:
            lines.append(f'    placeholder: "{jstr(extra)}",')
        lines.append(f'  }},')
    lines.append('];')
    return '\n'.join(lines)


with open(MAIN_JS_FILE, "r", encoding="utf-8") as f:
    js = f.read()

js = re.sub(r'const SENSOR_TYPES = \{.*?\};',
            build_sensor_types(), js, flags=re.DOTALL)
js = re.sub(r'const PORT_TIPS = \{.*?\};',
            build_port_tips(), js, flags=re.DOTALL)
js = re.sub(r'const PORTS = \[.*?\];',
            build_ports(), js, flags=re.DOTALL)
js = re.sub(r'const VIZ_OPTIONS = \[.*?\];',
            build_viz_options(), js, flags=re.DOTALL)
js = re.sub(r'const SURVEY_QUESTIONS = \[.*?\];',
            build_survey_questions(), js, flags=re.DOTALL)

with open(MAIN_JS_FILE, "w", encoding="utf-8") as f:
    f.write(js)

print(f"main.js updated — {len(active_keys)} sensor(s), {len(active_ports)} port(s), "
      f"{len(active_viz)} viz option(s), {len(survey_raw)} survey question(s)")


os.makedirs(TEMPLATES_DIR, exist_ok=True)

for key in active_keys:
    meta    = sensor_meta[key]
    params  = params_map.get(key, [])
    outputs = outputs_map.get(key, [])

    param_lines = []
    for p in params:
        mn = f"  # min: {p['min']}" if p["min"] else ""
        mx = f"  max: {p['max']}" if p["max"] else ""
        param_lines.append(f'{p["name"]} = {p["value"]}  # {p["label"]}{mn}{mx}')
    param_block = "\n".join(param_lines) if param_lines else "# No parameters defined"

    output_lines = []
    for o in outputs:
        eq     = f"  # equation: {o['equation']}" if o["equation"] else ""
        inputs = [i for i in o["inputs"] if i]
        inp    = f"  # inputs: {', '.join(inputs)}" if inputs else ""
        output_lines.append(f'#   - {o["value"]}{inp}{eq}')
    output_block = "\n".join(output_lines) if output_lines else "#   (none defined)"

    p0 = params[0]["name"] if len(params) > 0 else "param_1"
    p1 = params[1]["name"] if len(params) > 1 else "param_2"

    code = f'''# {key}.py
# Sensor: {meta["label"]}

{param_block}

# Supported outputs:
{output_block}


def read(raw_value: float) -> dict:
    result = {{}}
    return result


if __name__ == "__main__":
    test_raw = 512
    output   = read(test_raw)
    print(f"Sensor : {meta['label']}")
    print(f"Raw    : {{test_raw}}")
    print(f"Output : {{output}}")
'''

    path = os.path.join(TEMPLATES_DIR, f"{key}.py")
    with open(path, "w", encoding="utf-8") as f:
        f.write(code)
    print(f"templates/{key}.py written")

print("Done.")